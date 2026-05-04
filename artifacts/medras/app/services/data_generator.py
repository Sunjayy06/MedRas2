"""Custom practice dataset generator for the 4-step Practice Wizard.

This is a separate, additive service: the existing template-based generator
in `dummy_data.py` is left untouched. This module accepts a free-form
variable list (with optional per-variable hints) and produces a realistic
clinical-looking DataFrame plus a formatted Excel file.

Design rules
------------
* Pure-Python; only depends on numpy, pandas, openpyxl (already vendored).
* Never represents real patient data. Every output is marked as practice
  data via the workbook's first row and the `is_practice` meta flag.
* Variable type is inferred from the column name first (the same heuristic
  the main classifier uses), then refined from any user-supplied hints
  (min/max, percentages, expected-effect direction).
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Variable name → type detection.
#
# Mirrors (loosely) the SCORE_NAME pattern used by `variable_classifier.py`
# but adapted for *generation* — we need to know what distribution to draw
# from, not just whether the column is scale/nominal/etc.
# ---------------------------------------------------------------------------

_BINARY_HINTS = {
    "sex", "gender", "smoker", "smoking", "diabetic", "diabetes",
    "hypertension", "hypertensive", "outcome", "alive", "dead", "death",
    "survived", "responder", "response", "complication", "readmission",
    "yes_no", "treated",
}

_NOMINAL_HINTS = {
    "group", "arm", "treatment", "treatment_arm", "diagnosis",
    "category", "ethnicity", "race", "blood_group", "stage", "grade",
}

_SCALE_HINTS = {
    "age", "weight", "height", "bmi", "waist", "hip", "neck",
    "hba1c", "fbs", "ppbs", "rbs", "glucose", "insulin",
    "sbp", "dbp", "map", "pulse", "hr", "rr", "spo2", "temp",
    "hb", "wbc", "rbc", "platelet", "esr", "crp",
    "ferritin", "iron", "tsh", "t3", "t4", "creatinine", "urea",
    "bun", "ldl", "hdl", "triglyceride", "cholesterol", "albumin",
    "bilirubin", "ast", "alt", "alp",
    "score", "scale", "index", "vas", "nrs", "hhs", "harris",
    "duration", "time", "los", "days", "weeks", "months", "years",
    "dose", "level", "rate", "count",
}

# Realistic value ranges for common medical variables (mean, sd, lo, hi).
_RANGE_LIBRARY: Dict[str, Tuple[float, float, float, float]] = {
    "age":          (45, 15, 18, 90),
    "weight":       (70, 12, 40, 130),
    "height":       (165, 10, 140, 200),
    "bmi":          (26, 4, 16, 45),
    "hba1c":        (7.0, 1.4, 4.5, 13.0),
    "fbs":          (130, 40, 70, 320),
    "ppbs":         (180, 55, 90, 380),
    "sbp":          (135, 18, 90, 200),
    "dbp":          (85, 11, 55, 130),
    "hr":           (78, 12, 45, 130),
    "pulse":        (78, 12, 45, 130),
    "spo2":         (97, 2, 85, 100),
    "hb":           (12.5, 1.8, 6, 18),
    "ferritin":     (90, 60, 5, 350),
    "creatinine":   (1.0, 0.4, 0.4, 4.5),
    "urea":         (28, 12, 8, 90),
    "ldl":          (118, 30, 50, 240),
    "hdl":          (48, 12, 22, 95),
    "cholesterol":  (190, 38, 100, 320),
    "triglyceride": (150, 60, 40, 500),
    "vas":          (5, 2.5, 0, 10),
    "nrs":          (5, 2.5, 0, 10),
    "score":        (60, 18, 0, 100),
    "hhs":          (70, 18, 0, 100),
    "duration":     (12, 8, 0, 72),
    "los":          (5, 3, 1, 30),
    "days":         (10, 6, 0, 60),
    "weeks":        (8, 4, 0, 24),
    "months":       (6, 4, 0, 24),
    "dose":         (50, 20, 5, 200),
}

_NAME_RE = re.compile(r"[a-z0-9]+")


def _norm(name: str) -> str:
    return "_".join(_NAME_RE.findall((name or "").lower()))


def detect_type(name: str) -> str:
    """Return one of `scale`, `binary`, `nominal` from the column name."""
    norm = _norm(name)
    parts = set(norm.split("_"))
    if parts & _BINARY_HINTS or norm in _BINARY_HINTS:
        return "binary"
    if parts & _NOMINAL_HINTS or norm in _NOMINAL_HINTS:
        return "nominal"
    if parts & _SCALE_HINTS or norm in _SCALE_HINTS:
        return "scale"
    # Heuristic fallback by suffix.
    if any(norm.endswith(s) for s in ("_score", "_index", "_scale", "_rate", "_level")):
        return "scale"
    if any(norm.endswith(s) for s in ("_status", "_flag", "_yn")):
        return "binary"
    return "scale"


def _range_for(name: str) -> Tuple[float, float, float, float]:
    """Return (mean, sd, lo, hi) for a known medical variable; else generic."""
    norm = _norm(name)
    if norm in _RANGE_LIBRARY:
        return _RANGE_LIBRARY[norm]
    # Try suffix hits (e.g. "fasting_hba1c" → "hba1c").
    for key, vals in _RANGE_LIBRARY.items():
        if norm.endswith("_" + key) or norm.startswith(key + "_"):
            return vals
    return (50.0, 15.0, 0.0, 100.0)


# ---------------------------------------------------------------------------
# Spec → DataFrame
# ---------------------------------------------------------------------------


@dataclass
class VariableSpec:
    name: str
    type: str = "scale"        # scale | binary | nominal
    min: Optional[float] = None
    max: Optional[float] = None
    percent: Optional[float] = None     # for binary: % positive (0-100)
    levels: List[str] = field(default_factory=list)  # for nominal
    is_outcome: bool = False


@dataclass
class GenerateSpec:
    objective: str
    outcome: str
    variables: List[VariableSpec]
    n: int = 60
    expected_effect: str = ""
    instructions: str = ""        # Q4 — analysis hints (free-text)
    seed: Optional[int] = None


def _scale_values(rng: np.random.Generator, n: int, name: str,
                  vmin: Optional[float], vmax: Optional[float]) -> np.ndarray:
    mean, sd, lo, hi = _range_for(name)
    if vmin is not None:
        lo = float(vmin)
    if vmax is not None:
        hi = float(vmax)
    if vmax is not None and vmin is not None:
        # Recentre the distribution on the user-supplied window.
        mean = (lo + hi) / 2.0
        sd = max((hi - lo) / 5.0, 0.5)
    arr = rng.normal(loc=mean, scale=sd, size=n).clip(lo, hi)
    # Round integers when the variable is naturally integer.
    if any(k in _norm(name) for k in ("age", "los", "days", "count", "visits", "score", "sbp", "dbp", "hr")):
        arr = np.round(arr).astype(int)
    else:
        arr = np.round(arr, 1)
    return arr


def _binary_values(rng: np.random.Generator, n: int, name: str,
                   percent: Optional[float]) -> np.ndarray:
    p = (percent if percent is not None else 50.0) / 100.0
    p = float(np.clip(p, 0.05, 0.95))
    norm = _norm(name)
    if "sex" in norm or "gender" in norm:
        labels = ("Male", "Female")
    elif "smoke" in norm:
        labels = ("Yes", "No")
    elif "outcome" in norm or "responder" in norm or "alive" in norm:
        labels = ("Good", "Poor")
    elif "death" in norm or "dead" in norm:
        labels = ("Yes", "No")
    else:
        labels = ("Yes", "No")
    return rng.choice(labels, size=n, p=[p, 1 - p])


def _nominal_values(rng: np.random.Generator, n: int, name: str,
                    levels: List[str]) -> np.ndarray:
    chosen = [lv for lv in (levels or []) if lv.strip()]
    if not chosen:
        norm = _norm(name)
        if "group" in norm or "arm" in norm or "treatment" in norm:
            chosen = ["Treatment", "Control"]
        elif "stage" in norm:
            chosen = ["I", "II", "III", "IV"]
        elif "grade" in norm:
            chosen = ["Grade 1", "Grade 2", "Grade 3"]
        else:
            chosen = ["A", "B", "C"]
    probs = np.full(len(chosen), 1.0 / len(chosen))
    return rng.choice(chosen, size=n, p=probs)


def _apply_expected_effect(df: pd.DataFrame, spec: GenerateSpec,
                           rng: np.random.Generator) -> None:
    """Lightly bias outcome / scale columns based on a free-text effect.

    We don't try to be a parser — we just look for a few clear cues like
    "higher in X", "X higher", "increased in X" and shift the relevant
    column's mean upward when the row's group column matches.
    """
    text = (spec.expected_effect or "").lower()
    if not text:
        return

    direction = 1
    if any(w in text for w in ("lower", "decrease", "reduced", "less")):
        direction = -1

    # Find a target scale variable mentioned in the text.
    target = None
    for var in spec.variables:
        if var.type == "scale" and _norm(var.name) and _norm(var.name).split("_")[0] in text:
            target = var.name
            break
    if target is None:
        return

    # Find a group/binary column to split on.
    group_col = None
    group_value = None
    for var in spec.variables:
        if var.type in ("nominal", "binary"):
            for level in df[var.name].dropna().unique():
                if str(level).lower() in text:
                    group_col = var.name
                    group_value = level
                    break
            if group_col:
                break
    if group_col is None:
        return

    mask = df[group_col] == group_value
    series = df[target]
    if pd.api.types.is_numeric_dtype(series):
        shift = direction * (series.std() or 1.0) * 0.7
        df.loc[mask, target] = (series[mask] + shift).round(
            1 if series.dtype.kind == "f" else 0
        )


def _inject_missing(df: pd.DataFrame, pct: float, rng: np.random.Generator,
                    keep: Optional[List[str]] = None) -> None:
    if pct <= 0:
        return
    keep_set = set(keep or [])
    for col in df.columns:
        if col == "Patient_ID" or col in keep_set:
            continue
        mask = rng.random(len(df)) < (pct / 100.0)
        df.loc[mask, col] = np.nan


def generate_dataset(spec: GenerateSpec, missing_pct: float = 5.0) -> pd.DataFrame:
    rng = np.random.default_rng(spec.seed if spec.seed is not None else 42)
    n = max(20, min(int(spec.n or 60), 500))

    cols: Dict[str, np.ndarray] = {"Patient_ID": np.arange(1, n + 1)}
    for var in spec.variables:
        name = var.name.strip() or f"Var{len(cols)}"
        t = (var.type or detect_type(name)).lower()
        if t == "binary":
            cols[name] = _binary_values(rng, n, name, var.percent)
        elif t == "nominal":
            cols[name] = _nominal_values(rng, n, name, var.levels)
        else:
            cols[name] = _scale_values(rng, n, name, var.min, var.max)

    df = pd.DataFrame(cols)
    _apply_expected_effect(df, spec, rng)
    _inject_missing(df, missing_pct, rng, keep=["Patient_ID"])
    return df


# ---------------------------------------------------------------------------
# Excel writer with formatting (red disclaimer + blue header + zebra rows).
# ---------------------------------------------------------------------------


_DISCLAIMER_TEXT = (
    "⚠ PRACTICE DATA — NOT REAL PATIENTS. For learning and demonstration "
    "only. Do not publish or use for clinical decisions."
)


def save_to_excel(df: pd.DataFrame, study_name: str = "Practice dataset") -> bytes:
    """Return an .xlsx file with the spec'd formatting baked in."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Practice data"

    n_cols = len(df.columns)
    last_col = get_column_letter(max(1, n_cols))

    # Row 1 — red disclaimer banner spanning all columns.
    ws.cell(row=1, column=1, value=f"{_DISCLAIMER_TEXT}  Study: {study_name}".strip())
    ws.merge_cells(f"A1:{last_col}1")
    cell = ws.cell(row=1, column=1)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="C0392B")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Row 2 — blue header row.
    for c, col_name in enumerate(df.columns, start=1):
        h = ws.cell(row=2, column=c, value=str(col_name))
        h.font = Font(bold=True, color="FFFFFF", size=11)
        h.fill = PatternFill("solid", fgColor="1F4E79")
        h.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows with zebra shading.
    light = PatternFill("solid", fgColor="F2F4F7")
    for r, row in enumerate(df.itertuples(index=False), start=3):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=("" if pd.isna(val) else val))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r % 2 == 1:
                cell.fill = light

    # Auto-fit column widths (cap at 28 to keep things sane).
    for c, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(c)
        max_len = max(
            len(str(col_name)),
            *(len(str(v)) for v in df[col_name].fillna("").astype(str).head(50)),
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 28)

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Convenience: parse the Q3 multi-line variable list into VariableSpec[].
# ---------------------------------------------------------------------------


def parse_variable_list(text: str) -> List[Dict[str, str]]:
    """Split a multi-line block into rows with auto-detected type."""
    out: List[Dict[str, str]] = []
    for raw in (text or "").splitlines():
        name = raw.strip().lstrip("-•*").strip()
        if not name:
            continue
        out.append({"name": name, "type": detect_type(name)})
    # Dedupe on lowercase name while preserving order.
    seen = set()
    deduped = []
    for row in out:
        key = row["name"].lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped
